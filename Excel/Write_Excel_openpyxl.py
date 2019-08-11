import openpyxl

wb = openpyxl.Workbook()
print(wb.active.title) #by default sheet will come (wil show the active sheet

#if you want to change name of the sheet and then enter data into it
ws = wb.active #whatever the active sheet creates object for that and save into ws
ws.title = "Khurana" #change the title of the first sheet

print(ws.title) #here with sheet object we need not to write active

ws['A4'].value = "Python"

wb.create_sheet(title = "Write") #if you want to create one more sheet
ws1  = wb["Write"]
ws1 ['A5'] = "+91-9971477077"



wb.save("E://openpyxl_write.xlsx")

#-----------------

