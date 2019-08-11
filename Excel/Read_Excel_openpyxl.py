import openpyxl

#Open Workbook
wb = openpyxl.load_workbook("E:\\openpyxl_file.xls")

sh = wb["Testing"] #This is the sheet name testing and sh is sheet object

print(sh['A1'].value) #first approach to read data from cell is to use with sheet object
print(sh['B2'].value)#here we are using direct row and column number

cl = sh.cell(1,1) #means first row and first column
print(cl.value) #second approach to het data from cell using column like first row and first column  this is not index

#third approach is to pass keyword argument
cl = sh.cell(column = 3, row = 1)
print(cl.value)

print(cl.row)
print(cl.column)

#Find out how many rows having data and read all rows and columns
rows = sh.max_row
columns = sh.max_column

print("Total rows are " + str(rows))
print("Total columns are "+ str(columns))

for i in range(1,rows+1):
    for j in range(1, columns+1):
        c = sh.cell(i,j)
        print(c.value)

#one more approach to read all data
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)